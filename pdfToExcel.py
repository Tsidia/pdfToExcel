import fitz  # PyMuPDF
from PIL import Image
import pytesseract
import os
import tabula
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure Tesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Adjust the path as needed

def select_file(title):
    root = Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(title=title)
    return file_path

def select_folder(title):
    root = Tk()
    root.withdraw()  # Hide the root window
    folder_path = filedialog.askdirectory(title=title)
    return folder_path

def pdf_to_images(pdf_path, search_text, output_folder):
    # Open the PDF file
    pdf_document = fitz.open(pdf_path)
    
    # Ensure the output folder exists
    os.makedirs(output_folder, exist_ok=True)
    
    images = []

    # Iterate through the pages to find those containing the search text
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        text = page.get_text()

        if search_text in text:
            # Convert the page to an image
            pix = page.get_pixmap()
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Save the image
            img_path = os.path.join(output_folder, f"page_{page_num + 1}.png")
            img.save(img_path)
            print(f"Saved page {page_num + 1} as an image.")
            
            # Load the image for processing
            img = Image.open(img_path)
            
            # Perform OCR to find the position of the "Part list"
            ocr_result = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
            
            # Find the position of the "Part list" text
            part_list_idx = None
            target = 0
            for i, text in enumerate(ocr_result['text']):
                print("Tesseract found text: " + text + "At position: " + str(i))
                if "part" in text.strip().lower():
                    target += 1
                    if target == 2:
                        part_list_idx = i
                        break
            
            if part_list_idx is not None:
                # Get the bounding box coordinates of the "Part list" text
                x, y, w, h = (ocr_result['left'][part_list_idx], 
                              ocr_result['top'][part_list_idx], 
                              ocr_result['width'][part_list_idx], 
                              ocr_result['height'][part_list_idx])
                
                # Crop the image to remove everything below the "Part list"
                cropped_img = img.crop((0, 0, img.width, y))
                cropped_img.save(img_path)
                print(f"Cropped and saved page {page_num + 1} image.")
                images.append(img_path)
            else:
                print(f"'Part list' not found on page {page_num + 1}.")
    
    return images

def extract_tables(pdf_path):
    # Extract tables from PDF
    tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
    filtered_tables = [table for table in tables if table.apply(lambda row: row.astype(str).str.contains('Item No.').any(), axis=1).any()]
    return filtered_tables

def save_to_excel(images, tables, output_folder):
    excel_path = os.path.join(output_folder, 'output.xlsx')
    wb = Workbook()

    for i, (image_path, table) in enumerate(zip(images, tables)):
        ws = wb.create_sheet(title=f'Sheet_{i+1}')
        
        # Add the table to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(table, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Calculate the position to insert the image
        table_width = len(table.columns)
        img = XLImage(image_path)
        img.anchor = ws.cell(row=1, column=table_width + 2).coordinate  # Place the image next to the table
        ws.add_image(img)
    
    # Remove the default sheet created by openpyxl
    wb.remove(wb['Sheet'])
    
    wb.save(excel_path)
    print(f'Saved Excel file to {excel_path}')

def main():
    pdf_path = select_file("Select the PDF file")
    output_folder = select_folder("Select the output folder")
    search_text = "Assembly Drawing & Part List"
    
    # Step 1: Extract images from PDF
    images = pdf_to_images(pdf_path, search_text, output_folder)
    
    # Step 2: Extract tables from PDF
    tables = extract_tables(pdf_path)
    
    # Step 3: Save images and tables to Excel
    save_to_excel(images, tables, output_folder)

if __name__ == '__main__':
    main()
